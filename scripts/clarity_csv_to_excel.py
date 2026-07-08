from pathlib import Path
import csv
import re
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


PAGEVIEW_HEADERS = ["Month", "url", "total_pageview"]
SCROLL_HEADERS = ["Month", "url", "scroll depth", "number of visitors"]

QUALITY_HEADERS = [
    "check type",
    "status",
    "message",
    "source file",
    "Month",
    "url",
]

EXPECTED_SCROLL_DEPTHS = list(range(5, 101, 5))


def read_csv_rows(file_path: Path) -> list[list[str]]:
    """Read Clarity CSV with BOM-safe UTF-8 handling."""
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def get_metadata_value(rows: list[list[str]], key: str) -> str | None:
    """Find metadata values such as Date range, Page views, Metric, etc."""
    key_norm = key.strip().lower()

    for row in rows:
        if len(row) >= 2 and row[0].strip().lower() == key_norm:
            return row[1].strip()

    return None


def parse_clarity_datetime(value: str) -> datetime:
    """
    Parse Clarity date values.

    Expected example:
    06/01/2026 12:00 AM
    """
    value = value.strip()

    possible_formats = [
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y",
    ]

    for fmt in possible_formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass

    raise ValueError(f"Could not parse date: {value}")


def parse_date_range(date_range_text: str) -> tuple[date, str]:
    """
    Convert Clarity date range into:
    - month_date: first day of the month as a date object
    - month_key: YYYY-MM string for sorting and merging

    Example input:
    06/01/2026 12:00 AM - 06/29/2026 11:59 PM

    Output:
    date(2026, 6, 1), "2026-06"
    """
    if not date_range_text:
        raise ValueError("Date range is missing")

    start_text = date_range_text.split(" - ", 1)[0].strip()
    start_dt = parse_clarity_datetime(start_text)

    month_date = date(start_dt.year, start_dt.month, 1)
    month_key = start_dt.strftime("%Y-%m")

    return month_date, month_key


def regex_to_url(url_regex: str) -> str:
    r"""
    Convert the Clarity 'Visited URL matches regex' value into the real URL.

    Example input:
    ^https://info-prodotto\.it/integratori-di-magnesio/(\?.*)?$

    Output:
    https://info-prodotto.it/integratori-di-magnesio/
    """
    if not url_regex:
        raise ValueError("Visited URL matches regex is missing")

    url = url_regex.strip()

    if url.startswith("^"):
        url = url[1:]

    if url.endswith("$"):
        url = url[:-1]

    # Remove optional query-string regex suffix such as (\?.*)?
    url = re.sub(r"\(\\\?\.\*\)\??$", "", url)

    # Unescape common URL characters from regex notation.
    url = url.replace(r"\.", ".")
    url = url.replace(r"\/", "/")
    url = url.replace(r"\-", "-")
    url = url.replace(r"\_", "_")

    return url


def clean_int(value: str) -> int:
    """Convert values such as '21,475' into 21475."""
    if value is None:
        raise ValueError("Integer value is missing")

    return int(str(value).replace(",", "").strip())


def find_scroll_header_index(rows: list[list[str]]) -> int:
    """Find the row index where the scroll table starts."""
    for index, row in enumerate(rows):
        normalized = [cell.strip().lower() for cell in row]

        if (
            len(normalized) >= 3
            and normalized[0] == "scroll depth"
            and normalized[1] == "no. of visitors"
            and normalized[2] == "% drop off"
        ):
            return index

    raise ValueError("Scroll table header could not be found")


def parse_clarity_csv(file_path: Path) -> tuple[dict, list[dict]]:
    """
    Parse one Clarity Scroll CSV into:
    - one pageview record
    - multiple scroll records

    The % drop off column is ignored.
    """
    rows = read_csv_rows(file_path)

    date_range = get_metadata_value(rows, "Date range")
    url_regex = get_metadata_value(rows, "Visited URL matches regex")
    page_views = get_metadata_value(rows, "Page views")
    metric = get_metadata_value(rows, "Metric")

    if metric and metric.strip().lower() != "scroll":
        raise ValueError(f"Metric is not Scroll. Found: {metric}")

    month_date, month_key = parse_date_range(date_range)
    url = regex_to_url(url_regex)
    total_pageview = clean_int(page_views)

    pageview_record = {
        "month_date": month_date,
        "month_key": month_key,
        "url": url,
        "total_pageview": total_pageview,
        "source_file": str(file_path),
    }

    header_index = find_scroll_header_index(rows)
    scroll_records = []

    for row in rows[header_index + 1:]:
        if len(row) < 3 or not row[0].strip():
            continue

        scroll_records.append({
            "month_date": month_date,
            "month_key": month_key,
            "url": url,
            "scroll_depth": clean_int(row[0]),
            "number_of_visitors": clean_int(row[1]),
            "source_file": str(file_path),
        })

    return pageview_record, scroll_records


def collect_csv_files(input_dir: Path) -> list[Path]:
    """
    Find all CSV files under the data folder recursively.

    Expected folder structure:
    data/
      baerbel_drexel/
        ratgeber-naturprodukte/
          Clarity_...csv
      flinndal/
        some-url/
          Clarity_...csv
    """
    return sorted(input_dir.rglob("*.csv"))

def log_merged_inputs(pageview_records: list[dict]) -> None:
    """
    Log which CSV files are merged into the same Month + URL group.
    """
    groups = {}

    for record in pageview_records:
        key = (
            record["month_key"],
            record["month_date"],
            record["url"],
        )

        if key not in groups:
            groups[key] = []

        groups[key].append(record)

    merged_groups = {
        key: records
        for key, records in groups.items()
        if len(records) > 1
    }

    if not merged_groups:
        print("\nMerged groups: none")
        return

    print(f"\nMerged groups found: {len(merged_groups)}")

    for key, records in sorted(merged_groups.items()):
        month_key, month_date, url = key

        print("\nMERGED GROUP")
        print(f"Month: {month_date.strftime('%d/%m/%Y')}")
        print(f"URL: {url}")
        print(f"Files merged: {len(records)}")

        merged_pageviews = 0

        for record in records:
            merged_pageviews += record["total_pageview"]
            source_path = Path(record["source_file"])

            print(
                f"  - {source_path.name} | "
                f"pageviews: {record['total_pageview']}"
            )

        print(f"  => merged total_pageview: {merged_pageviews}")

def format_month_for_log(month_date: date | None) -> str:
    if month_date is None:
        return ""

    return month_date.strftime("%d/%m/%Y")


def add_quality_row(
    quality_rows: list[list],
    check_type: str,
    status: str,
    message: str,
    source_file: str = "",
    month_date: date | None = None,
    url: str = "",
):
    quality_rows.append([
        check_type,
        status,
        message,
        source_file,
        format_month_for_log(month_date),
        url,
    ])


def validate_parsed_file(
    csv_file: Path,
    pageview_record: dict,
    file_scroll_records: list[dict],
    quality_rows: list[list],
):
    """
    Validate one successfully parsed CSV file.
    These checks do not stop the script; they write PASS/WARN/FAIL rows
    into the Quality_Check sheet.
    """
    source_file = str(csv_file)
    month_date = pageview_record.get("month_date")
    url = pageview_record.get("url")
    total_pageview = pageview_record.get("total_pageview")

    if total_pageview is None:
        add_quality_row(
            quality_rows,
            "file validation",
            "FAIL",
            "Page views value is missing.",
            source_file,
            month_date,
            url,
        )
    elif total_pageview <= 0:
        add_quality_row(
            quality_rows,
            "file validation",
            "WARN",
            f"Page views value is zero or negative: {total_pageview}",
            source_file,
            month_date,
            url,
        )

    scroll_depths = [
        record["scroll_depth"]
        for record in file_scroll_records
    ]

    missing_depths = sorted(set(EXPECTED_SCROLL_DEPTHS) - set(scroll_depths))
    extra_depths = sorted(set(scroll_depths) - set(EXPECTED_SCROLL_DEPTHS))

    duplicate_depths = sorted({
        depth
        for depth in scroll_depths
        if scroll_depths.count(depth) > 1
    })

    if missing_depths:
        add_quality_row(
            quality_rows,
            "scroll depth validation",
            "WARN",
            f"Missing scroll depths: {missing_depths}",
            source_file,
            month_date,
            url,
        )

    if extra_depths:
        add_quality_row(
            quality_rows,
            "scroll depth validation",
            "WARN",
            f"Unexpected scroll depths: {extra_depths}",
            source_file,
            month_date,
            url,
        )

    if duplicate_depths:
        add_quality_row(
            quality_rows,
            "scroll depth validation",
            "WARN",
            f"Duplicate scroll depths in one CSV: {duplicate_depths}",
            source_file,
            month_date,
            url,
        )

    if not missing_depths and not extra_depths and not duplicate_depths:
        add_quality_row(
            quality_rows,
            "scroll depth validation",
            "PASS",
            "Scroll depths look complete and unique.",
            source_file,
            month_date,
            url,
        )

    if total_pageview and total_pageview > 0:
        visitors_above_pageviews = []

        for record in file_scroll_records:
            if record["number_of_visitors"] > total_pageview:
                visitors_above_pageviews.append(
                    record["scroll_depth"]
                )

        if visitors_above_pageviews:
            add_quality_row(
                quality_rows,
                "visitor validation",
                "WARN",
                (
                    "Number of visitors is higher than total pageviews "
                    f"for scroll depths: {visitors_above_pageviews}"
                ),
                source_file,
                month_date,
                url,
            )
        else:
            add_quality_row(
                quality_rows,
                "visitor validation",
                "PASS",
                "Number of visitors values are not higher than total pageviews.",
                source_file,
                month_date,
                url,
            )


def log_and_validate_merged_groups(
    pageview_records: list[dict],
    quality_rows: list[list],
):
    """
    Detect and log which month + url combinations were merged.
    """
    groups = {}

    for record in pageview_records:
        key = (
            record["month_key"],
            record["month_date"],
            record["url"],
        )

        if key not in groups:
            groups[key] = []

        groups[key].append(record)

    merged_groups = {
        key: records
        for key, records in groups.items()
        if len(records) > 1
    }

    if not merged_groups:
        print("\nMerged groups: none")

        add_quality_row(
            quality_rows,
            "merge validation",
            "PASS",
            "No merged groups found.",
        )

        return

    print(f"\nMerged groups found: {len(merged_groups)}")

    for key, records in sorted(merged_groups.items()):
        month_key, month_date, url = key

        merged_pageviews = sum(
            record["total_pageview"]
            for record in records
        )

        file_names = [
            Path(record["source_file"]).name
            for record in records
        ]

        print("\nMERGED GROUP")
        print(f"Month: {month_date.strftime('%d/%m/%Y')}")
        print(f"URL: {url}")
        print(f"Files merged: {len(records)}")

        for record in records:
            source_path = Path(record["source_file"])
            print(
                f"  - {source_path.name} | "
                f"pageviews: {record['total_pageview']}"
            )

        print(f"  => merged total_pageview: {merged_pageviews}")

        add_quality_row(
            quality_rows,
            "merge validation",
            "WARN",
            (
                f"{len(records)} CSV files were merged. "
                f"Merged total_pageview: {merged_pageviews}. "
                f"Files: {', '.join(file_names)}"
            ),
            "",
            month_date,
            url,
        )


def add_summary_quality_checks(
    quality_rows: list[list],
    csv_count: int,
    processed_count: int,
    skipped_count: int,
    pageview_rows_count: int,
    scroll_rows_count: int,
):
    add_quality_row(
        quality_rows,
        "summary",
        "PASS" if csv_count == processed_count + skipped_count else "FAIL",
        (
            f"CSV count: {csv_count}, "
            f"processed: {processed_count}, "
            f"skipped: {skipped_count}"
        ),
    )

    add_quality_row(
        quality_rows,
        "summary",
        "PASS",
        f"Final Pageviews rows: {pageview_rows_count}",
    )

    add_quality_row(
        quality_rows,
        "summary",
        "PASS",
        f"Final Scroll rows: {scroll_rows_count}",
    )
    
def merge_pageviews(pageview_records: list[dict]) -> list[list]:
    """
    Merge pageview records by Month + url.

    If one month has multiple partial CSV exports,
    total_pageview is summed.
    """
    merged = {}

    for record in pageview_records:
        key = (
            record["month_key"],
            record["month_date"],
            record["url"], 
        )

        if key not in merged:
            merged[key] = 0

        merged[key] += record["total_pageview"]

    output_rows = []

    for month_key, month_date, url in sorted(merged.keys()):
        output_rows.append([
            month_date,
            url,
            merged[(month_key, month_date, url)],
        ])

    return output_rows


def merge_scroll(scroll_records: list[dict]) -> list[list]:
    """
    Merge scroll records by Month + url + scroll depth.

    If one month has multiple partial CSV exports:
    number of visitors is summed.

    % drop off is ignored and not included in the final Excel.
    """
    merged = {}

    for record in scroll_records:
        key = (
            record["month_key"],
            record["month_date"],
            record["url"],
            record["scroll_depth"],
        )

        if key not in merged:
            merged[key] = 0

        merged[key] += record["number_of_visitors"]

    output_rows = []

    for month_key, month_date, url, scroll_depth in sorted(merged.keys()):
        output_rows.append([
            month_date,
            url,
            scroll_depth,
            merged[(month_key, month_date, url, scroll_depth)],
        ])

    return output_rows


def autofit_basic(ws):
    """Apply practical column widths."""
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)

        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        width = min(max(max_length + 2, 12), 70)
        ws.column_dimensions[column_letter].width = width


def apply_number_formats(ws):
    """Apply Excel number/date formats."""
    headers = [cell.value for cell in ws[1]]

    for col_index, header in enumerate(headers, start=1):
        if header == "Month":
            for cell in ws.iter_cols(
                min_col=col_index,
                max_col=col_index,
                min_row=2,
                max_row=ws.max_row,
            ):
                for c in cell:
                    c.number_format = "DD/MM/YYYY"

        if header in ["total_pageview", "number of visitors", "scroll depth"]:
            for cell in ws.iter_cols(
                min_col=col_index,
                max_col=col_index,
                min_row=2,
                max_row=ws.max_row,
            ):
                for c in cell:
                    c.number_format = "#,##0"


def style_sheet(ws, table_name: str):
    """Basic readable Excel formatting."""
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    apply_number_formats(ws)
    autofit_basic(ws)

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row > 1:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=table_name, ref=ref)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def build_excel(input_dir: Path, output_file: Path):
    pageview_records = []
    scroll_records = []
    quality_rows = []
    errors = []

    csv_files = collect_csv_files(input_dir)

    if not csv_files:
        raise FileNotFoundError(f"No CSV files found under: {input_dir}")

    print(f"Found CSV files: {len(csv_files)}")

    for csv_file in csv_files:
        try:
            pageview_record, file_scroll_records = parse_clarity_csv(csv_file)

            pageview_records.append(pageview_record)
            scroll_records.extend(file_scroll_records)

            validate_parsed_file(
                csv_file,
                pageview_record,
                file_scroll_records,
                quality_rows,
            )

            print(f"Processed: {csv_file}")

        except Exception as exc:
            errors.append((str(csv_file), str(exc)))

            add_quality_row(
                quality_rows,
                "file parsing",
                "FAIL",
                str(exc),
                str(csv_file),
            )

            print(f"Skipped: {csv_file} | Reason: {exc}")

    log_and_validate_merged_groups(pageview_records, quality_rows)

    pageview_rows = merge_pageviews(pageview_records)
    scroll_rows = merge_scroll(scroll_records)

    add_summary_quality_checks(
        quality_rows=quality_rows,
        csv_count=len(csv_files),
        processed_count=len(pageview_records),
        skipped_count=len(errors),
        pageview_rows_count=len(pageview_rows),
        scroll_rows_count=len(scroll_rows),
    )

    wb = Workbook()

    pageview_ws = wb.active
    pageview_ws.title = "Pageviews"

    scroll_ws = wb.create_sheet("Scroll")
    quality_ws = wb.create_sheet("Quality_Check")

    pageview_ws.append(PAGEVIEW_HEADERS)
    for row in pageview_rows:
        pageview_ws.append(row)

    scroll_ws.append(SCROLL_HEADERS)
    for row in scroll_rows:
        scroll_ws.append(row)

    quality_ws.append(QUALITY_HEADERS)
    for row in quality_rows:
        quality_ws.append(row)

    style_sheet(pageview_ws, "PageviewsTable")
    style_sheet(scroll_ws, "ScrollTable")
    style_sheet(quality_ws, "QualityCheckTable")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    print("\nDone")
    print(f"Output file: {output_file}")
    print(f"CSV files found: {len(csv_files)}")
    print(f"CSV files processed: {len(pageview_records)}")
    print(f"CSV files skipped: {len(errors)}")
    print(f"Pageview rows: {len(pageview_rows)}")
    print(f"Scroll rows: {len(scroll_rows)}")
    print(f"Quality check rows: {len(quality_rows)}")

    if errors:
        print("\nFiles skipped:")
        for file_name, reason in errors:
            print(f"- {file_name}: {reason}")

def main():
    project_root = Path(__file__).resolve().parents[1]

    input_dir = project_root / "data"
    output_file = project_root / "output" / "clarity_combined_output.xlsx"

    build_excel(input_dir, output_file)


if __name__ == "__main__":
    main()